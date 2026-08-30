"""official_career_50 校标 xlsx → 评测用 gold JSONL（jd_golden_110 同 schema）。

用法（backend/ 下）：
    python tests/evaluate/build_official_career_gold.py

输入：data/golden_set/candidate_pool/official_career_50/official_career_50_pre_annotate.xlsx
输出：data/golden_set/candidate_pool/official_career_50/official_career_50_gold.jsonl

字段映射（对齐 data/golden_set/final/jd_golden_110.jsonl 数据字典）：
    gold_title        ← review_gold_title
    gold_skills       ← review_gold_skills（JSON 数组）
    gold_bonus_skills ← review_gold_bonus_skills（JSON 数组）
    gold_education    ← review_gold_education（级别字符串；空 = 无明确学历要求）
    gold_core_duties  ← review_gold_core_duties（JSON 数组）
    gold_experience   ← review_gold_experience（JSON 对象或缺失）

前置校验：review_status 必须全部 DONE；review_gold_* 不得残留空技能列；
JSON 列必须可解析。任何一条不满足即 exit(1)，不产出半成品 gold。
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "data" / "golden_set" / "candidate_pool" / "official_career_50" / "official_career_50_pre_annotate.xlsx"
OUT = ROOT / "data" / "golden_set" / "candidate_pool" / "official_career_50" / "official_career_50_gold.jsonl"

ARRAY_FIELDS = ("gold_skills", "gold_bonus_skills", "gold_core_duties")


def _parse_json(value, field: str, sid: str) -> list:
    if value is None or not str(value).strip():
        return []
    try:
        decoded = json.loads(value)
    except json.JSONDecodeError as exc:
        print(f"BLOCKED: {sid} {field} JSON 解析失败：{exc}")
        sys.exit(1)
    if not isinstance(decoded, list) or not all(isinstance(x, str) for x in decoded):
        print(f"BLOCKED: {sid} {field} 不是字符串数组")
        sys.exit(1)
    return decoded


def main() -> int:
    ws = load_workbook(XLSX).active
    header = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header) if h}

    not_done = []
    for r in range(2, ws.max_row + 1):
        status = (ws.cell(r, col["review_status"]).value or "").strip()
        if status != "DONE":
            not_done.append((r, status or "(空)"))
    if not_done:
        print(f"BLOCKED: {len(not_done)} 行 review_status != DONE：{not_done[:5]}")
        return 1

    out_lines: list[str] = []
    for r in range(2, ws.max_row + 1):
        sid = str(ws.cell(r, col["source_id"]).value or "")
        rec = {
            "sample_id": str(ws.cell(r, col["sample_id"]).value or f"oc50_{sid[-12:]}"),
            "source": str(ws.cell(r, col["source"]).value or ""),
            "source_id": sid,
            "source_url": str(ws.cell(r, col["source_url"]).value or ""),
            "job_title_raw": str(ws.cell(r, col["job_title_raw"]).value or ""),
            "detail_raw_text": str(ws.cell(r, col["detail_raw_text"]).value or ""),
            "text_education": str(ws.cell(r, col["text_education"]).value or ""),
            "source_education": str(ws.cell(r, col["source_education"]).value or ""),
        }
        for out_field, col_name in (
            ("gold_title", "review_gold_title"),
            ("gold_education", "review_gold_education"),
        ):
            rec[out_field] = str(ws.cell(r, col[col_name]).value or "").strip()
        for out_field, col_name in (
            ("gold_skills", "review_gold_skills"),
            ("gold_bonus_skills", "review_gold_bonus_skills"),
            ("gold_core_duties", "review_gold_core_duties"),
        ):
            rec[out_field] = _parse_json(ws.cell(r, col[col_name]).value, out_field, sid)
        # experience：校标全留空（08-28 拍板口径），留缺失与 110 null 语义一致
        exp_raw = str(ws.cell(r, col["review_gold_experience"]).value or "").strip()
        if exp_raw:
            try:
                rec["gold_experience"] = json.loads(exp_raw)
            except json.JSONDecodeError as exc:
                print(f"BLOCKED: {sid} gold_experience JSON 解析失败：{exc}")
                return 1
        empty_skills = not rec["gold_skills"]
        if empty_skills:
            print(f"BLOCKED: {sid} gold_skills 为空（校标不允许空技能集）")
            return 1
        out_lines.append(json.dumps(rec, ensure_ascii=False))

    OUT.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    print(f"OK: {len(out_lines)} 条 gold 写入 {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
