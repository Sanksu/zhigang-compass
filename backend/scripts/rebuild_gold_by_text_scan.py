"""盲审 gold 按「任职要求段技术全收」口径重建（2026-08-16 方案 A-收窄）。

背景：JD 解析 0.90 攻坚——方案 A 全量（正文技术全收）实证失败（F1 0.7376，
recall 崩 0.63：gold 30-60 词 vs 预测克制 25-37 词，漏抽 219）。用户决策
A-收窄：gold 只收「任职要求/岗位要求/任职资格/Requirements」段落的技术词
（扫描限定段落）+ 原 gold 并集——预测的克制行为可匹配，避免全收 recall 崩。

口径：
- 定位任职要求段标记（任职要求/岗位要求/任职资格/Requirements），从标记
  位置扫描至文末；无标记样本退回全文扫描（仅扫描集为全文）
- 扫描源 = SKILL_WHITELIST（605）+ SKILL_ALIAS 键（130，规范到标准写法）
- 词边界匹配（防 "c" 命中 "Python" 子串）；过滤软技能白名单与技能停用词
- gold_v2 = 段落扫描命中 ∪ 原 gold skills（并集——保留人工标定的白名单外词）

注意：gold 口径变更后 F1 与历史基线（0.7563/0.8223）不可直接对比。

用法（cwd=backend）：
    python -m scripts.rebuild_gold_by_text_scan           # dry-run 报告
    python -m scripts.rebuild_gold_by_text_scan --apply   # 写回 xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.core.logging import setup_logging
from app.services.extraction.dictionary import (
    SOFT_SKILL_WHITELIST,
    _SKILL_WHITELIST_LOWER,
)
from app.services.extraction.dictionary_data import SKILL_ALIAS, SKILL_STOPWORDS

logger = setup_logging("rebuild_gold_by_text_scan")

_XLSX = ROOT / "data" / "golden_set" / "review" / "jd_manual_review_merged.xlsx"
_SHEET = "Round合并盲标"

# 任职要求段标记（取最早出现者，从该处扫描至文末）
_REQ_MARKERS = ("任职要求", "岗位要求", "任职资格", "requirements")

soft = {s.lower() for s in SOFT_SKILL_WHITELIST}
stop = {s.lower() for s in SKILL_STOPWORDS}
wordlist: dict[str, str] = {}
for w in _SKILL_WHITELIST_LOWER:
    wordlist.setdefault(w, w)
for k, v in SKILL_ALIAS.items():
    wordlist.setdefault(k.lower(), v)
WORDS = sorted(wordlist, key=len, reverse=True)  # 长词优先


def _requirement_section(text: str) -> str:
    """任职要求段：最早标记起至文末；无标记返回全文。"""
    low = text.lower()
    pos = [low.find(m) for m in _REQ_MARKERS if low.find(m) >= 0]
    return text[min(pos):] if pos else text


def scan_skills(text: str) -> set[str]:
    """词边界扫描任职要求段命中白名单技能（规范写法，过滤软技能/停用词）。"""
    low = _requirement_section(text).lower()
    hits: set[str] = set()
    for w in WORDS:
        if w in soft or w in stop or len(w) < 2:
            continue
        if re.search(r"(?<![a-z0-9])" + re.escape(w) + r"(?![a-z0-9])", low):
            hits.add(wordlist[w])
    return hits


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="盲审 gold 正文技术全收口径重建")
    parser.add_argument("--apply", action="store_true", help="写回 xlsx（默认 dry-run）")
    args = parser.parse_args(argv)

    from openpyxl import load_workbook

    wb = load_workbook(_XLSX)
    ws = wb[_SHEET]
    header = [c.value for c in ws[1]]
    col_text = header.index("detail_raw_text")
    col_skills = header.index("review_gold_skills")

    changes = 0
    for row in ws.iter_rows(min_row=2):
        sid = str(row[0].value or "").strip()
        if not sid:
            continue
        text = row[col_text].value or ""
        scanned = scan_skills(text)
        original = set(json.loads(row[col_skills].value or "[]"))
        merged = sorted(scanned | original)
        if len(merged) != len(original):
            changes += 1
            logger.info(
                "  %s: gold %d → %d（扫描 %d）", sid, len(original), len(merged), len(scanned)
            )
            if args.apply:
                row[col_skills].value = json.dumps(merged, ensure_ascii=False)
    if args.apply:
        wb.save(_XLSX)
        logger.info("已写回 %s 条 gold（%s）", changes, _XLSX)
    else:
        logger.info("dry-run：%s 条 gold 将变更（--apply 写回）", changes)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
