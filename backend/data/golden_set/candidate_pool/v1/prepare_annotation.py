#!/usr/bin/env python3
"""
prepare_annotation.py — 从 accepted 候选池中建立人工标注数据包

确定性分层抽样 (seed=20260817):
    primary 110 条 + reserve 25 条 = accepted 135 条

输出:
    annotation/jd_annotation_round1_110.xlsx     (主标注集)
    annotation/jd_annotation_reserve_25.xlsx     (备用集)
    annotation/annotation_sampling_manifest.csv  (样本清单)
"""

import json
import os
import csv
import random
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 路径配置 ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ANNOTATION_DIR = os.path.join(os.path.dirname(BASE_DIR), "annotation")
RAW_FILE = os.path.join(BASE_DIR, "real_jd_candidates_raw.jsonl")

SEED = 20260817
random.seed(SEED)

# ─── 分层抽样方案 ──────────────────────────────────────────────────────
# category → (primary_count, reserve_count)
SAMPLING_PLAN = {
    "后端开发": (26, 6),
    "AI/大模型": (14, 3),
    "算法": (12, 3),
    "嵌入式/C++": (11, 2),
    "全栈开发": (11, 2),
    "运维/DevOps": (10, 2),
    "数据工程/大数据": (8, 2),
    "前端开发": (6, 2),
    "数据分析": (4, 1),
    "测试": (4, 1),
    "网络/安全": (2, 1),
    "其他技术岗": (2, 0),
}

# ─── 加载数据 ──────────────────────────────────────────────────────────

def load_accepted():
    """加载 accepted 记录"""
    records = []
    with open(RAW_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                r = json.loads(line)
                if r.get("tier") == "accepted":
                    records.append(r)
    return records


# ─── 分层抽样 ──────────────────────────────────────────────────────────

def stratified_sample(records):
    """确定性分层抽样，返回 (primary, reserve)"""
    # 按 category 分组
    by_cat = defaultdict(list)
    for r in records:
        cat = r.get("category", "其他技术岗")
        by_cat[cat].append(r)
    
    # 每组内按 source_id 排序保证确定性
    for cat in by_cat:
        by_cat[cat].sort(key=lambda r: r.get("source_id", ""))
    
    primary = []
    reserve = []
    
    for cat, (p_count, r_count) in SAMPLING_PLAN.items():
        pool = by_cat.get(cat, [])
        if len(pool) < p_count + r_count:
            print(f"  ⚠️ {cat}: 实际 {len(pool)} 条, 期望 {p_count + r_count} 条, 调整中...")
            p_count = min(p_count, len(pool))
            r_count = min(r_count, len(pool) - p_count)
        
        # 确定性随机抽样
        indices = list(range(len(pool)))
        random.shuffle(indices)
        
        primary_indices = sorted(indices[:p_count])
        reserve_indices = sorted(indices[p_count:p_count + r_count])
        
        for i in primary_indices:
            r = dict(pool[i])
            r["annotation_split"] = "primary"
            primary.append(r)
        
        for i in reserve_indices:
            r = dict(pool[i])
            r["annotation_split"] = "reserve"
            reserve.append(r)
    
    # 按 source_id 排序
    primary.sort(key=lambda r: r.get("source_id", ""))
    reserve.sort(key=lambda r: r.get("source_id", ""))
    
    # 分配 sample_id
    for i, r in enumerate(primary):
        r["sample_id"] = f"ANN-{i+1:04d}"
    for i, r in enumerate(reserve):
        r["sample_id"] = f"RES-{i+1:04d}"
    
    return primary, reserve


# ─── Excel 样式 ─────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="微软雅黑", size=10)
BLANK_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淡黄色标注人工填写列
NOTE_FONT = Font(name="微软雅黑", size=10, color="333333")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def style_header_row(ws, row, num_cols):
    """给标题行上色"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, is_blank=False):
    """设置数据单元格样式"""
    cell = ws.cell(row=row, column=col)
    cell.font = CELL_FONT
    cell.alignment = WRAP_ALIGN
    cell.border = THIN_BORDER
    if is_blank:
        cell.fill = BLANK_FILL


# ─── 生成 Excel ─────────────────────────────────────────────────────────

# 原始证据列
EVIDENCE_FIELDS = [
    "sample_id",
    "source",
    "source_id",
    "source_url",
    "job_title_raw",
    "company_name",
    "location",
    "detail_raw_text",
    "responsibilities",
    "requirements",
    "source_education",
    "source_experience",
    "text_education",
    "text_experience",
    "education_conflict",
    "experience_conflict",
]

# 人工填写列 (全部空白)
ANNOTATION_FIELDS = [
    "review_gold_title",
    "review_gold_skills",
    "review_gold_bonus_skills",
    "review_gold_experience",
    "review_gold_education",
    "review_gold_core_duties",
    "annotator",
    "review_status",
    "error_type",
    "review_note",
]

ALL_FIELDS = EVIDENCE_FIELDS + ANNOTATION_FIELDS


def create_annotation_sheet(ws, records, sheet_title):
    """创建标注工作表"""
    ws.title = sheet_title
    
    # 写入表头
    for col_idx, field in enumerate(ALL_FIELDS, 1):
        ws.cell(row=1, column=col_idx, value=field)
    style_header_row(ws, 1, len(ALL_FIELDS))
    
    # 写入数据
    for row_idx, r in enumerate(records, 2):
        for col_idx, field in enumerate(ALL_FIELDS, 1):
            if field in EVIDENCE_FIELDS:
                val = r.get(field, "")
                if isinstance(val, bool):
                    val = str(val)
                elif val is None:
                    val = ""
                ws.cell(row=row_idx, column=col_idx, value=val)
                style_data_cell(ws, row_idx, col_idx, is_blank=False)
            else:
                # 人工填写列 → 全部空白
                ws.cell(row=row_idx, column=col_idx, value="")
                style_data_cell(ws, row_idx, col_idx, is_blank=True)
    
    # 设置列宽
    col_widths = {
        "sample_id": 12,
        "source": 10,
        "source_id": 28,
        "source_url": 45,
        "job_title_raw": 25,
        "company_name": 20,
        "location": 15,
        "detail_raw_text": 60,
        "responsibilities": 50,
        "requirements": 50,
        "source_education": 14,
        "source_experience": 14,
        "text_education": 14,
        "text_experience": 14,
        "education_conflict": 14,
        "experience_conflict": 14,
        "review_gold_title": 22,
        "review_gold_skills": 30,
        "review_gold_bonus_skills": 30,
        "review_gold_experience": 20,
        "review_gold_education": 16,
        "review_gold_core_duties": 40,
        "annotator": 12,
        "review_status": 14,
        "error_type": 16,
        "review_note": 30,
    }
    
    for col_idx, field in enumerate(ALL_FIELDS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 15)
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_FIELDS))}{len(records) + 1}"


def create_instruction_sheet(ws):
    """创建标注说明工作表"""
    ws.title = "标注说明"
    
    instructions = [
        ("# JD 人工标注规范 (Round1 盲标)", "title"),
        ("", "blank"),
        ("## 重要原则", "heading"),
        ("", "blank"),
        ("1. 本轮为盲标（Blind Annotation），标注人看不到任何模型预测结果。", "text"),
        ("2. 所有标注信息必须严格来源于 JD 正文，不得根据常识或岗位一般知识补全。", "text"),
        ("3. 黄色背景列为人工填写区域。能够根据 JD 正文明确信息判断的字段应填写；原文没有明确证据的字段必须留空，不得为了完整率猜测或补写。", "text"),
        ("4. 如果某项信息在 JD 正文中无法找到，请留空，不要猜测。", "text"),
        ("", "blank"),
        ("## 格式校验说明", "heading"),
        ("", "blank"),
        ("以下字段必须使用 JSON 数组格式（字符串数组）：", "text"),
        ('  - review_gold_skills         示例: ["Python","TensorFlow","Docker"]', "text"),
        ('  - review_gold_bonus_skills   示例: ["Redis","Docker"]', "text"),
        ('  - review_gold_core_duties    示例: ["负责后端API设计与开发","参与系统架构评审"]', "text"),
        ("", "blank"),
        ("原文没有对应内容时允许填写空数组：[]", "text"),
        ("", "blank"),
        ("以下字段使用 JSON 对象格式：", "text"),
        ('  - review_gold_experience     示例: {"min_years":3,"max_years":null}', "text"),
        ("", "blank"),
        ("以下字段为普通字符串：", "text"),
        ('  - review_gold_title          示例: "Java后端开发工程师"', "text"),
        ('  - review_gold_education      示例: "本科"', "text"),
        ('  - annotator                  示例: "张三"', "text"),
        ('  - review_status              示例: "已完成"', "text"),
        ('  - error_type                 示例: "信息缺失"', "text"),
        ('  - review_note                示例: "JD正文中未明确学历要求"', "text"),
        ("", "blank"),
        ("## 字段说明", "heading"),
        ("", "blank"),
        ("### review_gold_title", "subheading"),
        ("类型: 普通字符串", "text"),
        ("依据 job_title_raw 和 detail_raw_text 正文，填写规范岗位名称。", "text"),
        ("不得根据常识改变岗位语义。例如正文是「Java开发」，不要写成「高级Java架构师」。", "text"),
        ("", "blank"),
        ("### review_gold_skills", "subheading"),
        ("类型: JSON 字符串数组", "text"),
        ('格式: ["Python","TensorFlow","Docker"]', "text"),
        ("仅填写正文中明确要求的必备技术、工具或方法。", "text"),
        ("不要填写正文未提及的技能。", "text"),
        ("原文没有对应内容时填写: []", "text"),
        ("", "blank"),
        ("### review_gold_bonus_skills", "subheading"),
        ("类型: JSON 字符串数组", "text"),
        ('格式: ["Redis","Docker"]', "text"),
        ("仅填写正文中明确出现以下语义的技能：", "text"),
        ("  - 优先", "text"),
        ("  - 加分", "text"),
        ("  - 熟悉更佳", "text"),
        ("  - 有经验者优先", "text"),
        ("  - 了解即可", "text"),
        ("  - 具有...经验者优先考虑", "text"),
        ("  - 类似非必备语义", "text"),
        ("不要将正文明确要求的技能放进 bonus_skills。", "text"),
        ("原文没有对应内容时填写: []", "text"),
        ("", "blank"),
        ("### review_gold_experience", "subheading"),
        ("类型: JSON 对象或空", "text"),
        ('格式: {"min_years":3,"max_years":null}', "text"),
        ("仅根据正文中明确要求的经验年限填写。", "text"),
        ("如果原文没有明确经验要求，该字段留空，不要猜测。", "text"),
        ("注意：如果 source_experience 与正文冲突，以人工阅读正文为准，同时保留冲突信息供复核。", "text"),
        ("", "blank"),
        ("### review_gold_education", "subheading"),
        ("类型: 普通字符串或空", "text"),
        ("只填写正文中明确的最低学历要求。", "text"),
        ("可选值: 大专 / 本科 / 硕士 / 博士 / 不限", "text"),
        ("如果原文未明确学历要求，该字段留空，不要猜测。", "text"),
        ("", "blank"),
        ("### review_gold_core_duties", "subheading"),
        ("类型: JSON 字符串数组", "text"),
        ('格式: ["负责后端API设计与开发","参与系统架构评审"]', "text"),
        ("只忠实概括正文中的职责，使用 JSON 数组格式。", "text"),
        ("不得补写 JD 中不存在的职责。", "text"),
        ("原文没有对应内容时填写: []", "text"),
        ("", "blank"),
        ("### annotator", "subheading"),
        ("类型: 普通字符串", "text"),
        ("填写标注人姓名或工号。", "text"),
        ("", "blank"),
        ("### review_status", "subheading"),
        ("类型: 普通字符串", "text"),
        ("可选值: 已完成 / 待确认 / 有争议", "text"),
        ("", "blank"),
        ("### error_type", "subheading"),
        ("类型: 普通字符串", "text"),
        ("如果发现 JD 数据本身存在问题，填写错误类型：", "text"),
        ("  - 非JD内容: 正文不是真正的岗位描述", "text"),
        ("  - 信息缺失: 正文信息严重不足", "text"),
        ("  - 字段错误: 原始字段与正文明显不符", "text"),
        ("  - 重复: 与其他样本高度重复", "text"),
        ("  - 其他: 其他问题", "text"),
        ("如无问题则留空。", "text"),
        ("", "blank"),
        ("### review_note", "subheading"),
        ("类型: 普通字符串", "text"),
        ("标注过程中的任何备注或疑问。", "text"),
        ("", "blank"),
        ("## 冲突字段说明", "heading"),
        ("", "blank"),
        ("- education_conflict = true: 表示 source_education 与正文提取的 text_education 存在明显矛盾", "text"),
        ("- experience_conflict = true: 表示 source_experience 与正文提取的 text_experience 存在明显矛盾", "text"),
        ("- 标注人应以人工阅读正文为准，同时关注冲突标记。", "text"),
        ("", "blank"),
        ("## 标注示例", "heading"),
        ("", "blank"),
        ("假设某 JD 的正文为：", "text"),
        ('"岗位职责：1. 负责公司核心业务系统的后端开发与维护；2. 参与系统架构设计。', "text"),
        ('任职要求：1. 本科及以上学历；2. 3年以上Java开发经验；3. 精通Spring Boot、MySQL；4. 熟悉Redis、Docker优先。"', "text"),
        ("", "blank"),
        ("则标注结果应为：", "text"),
        ('  review_gold_title:          "Java后端开发工程师"', "text"),
        ('  review_gold_skills:         ["Java","Spring Boot","MySQL"]', "text"),
        ('  review_gold_bonus_skills:   ["Redis","Docker"]', "text"),
        ('  review_gold_experience:     {"min_years":3,"max_years":null}', "text"),
        ('  review_gold_education:      "本科"', "text"),
        ('  review_gold_core_duties:    ["负责公司核心业务系统的后端开发与维护","参与系统架构设计"]', "text"),
    ]
    
    # 设置列宽
    ws.column_dimensions["A"].width = 100
    
    for row_idx, (text, style) in enumerate(instructions, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if style == "title":
            cell.font = TITLE_FONT
        elif style == "heading":
            cell.font = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
        elif style == "subheading":
            cell.font = Font(name="微软雅黑", size=11, bold=True, color="333333")
        else:
            cell.font = NOTE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def create_stats_sheet(ws, primary, reserve):
    """创建数据统计工作表"""
    ws.title = "数据统计"
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    
    # 标题
    ws.cell(row=1, column=1, value="标注数据统计").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = NOTE_FONT
    ws.cell(row=3, column=1, value=f"随机种子: {SEED}").font = NOTE_FONT
    
    # 总体统计
    row = 5
    ws.cell(row=row, column=1, value="总体统计").font = Font(name="微软雅黑", size=12, bold=True)
    row += 1
    for label, val in [("主标注集 (primary)", len(primary)), ("备用集 (reserve)", len(reserve)), ("合计", len(primary) + len(reserve))]:
        ws.cell(row=row, column=1, value=label).font = NOTE_FONT
        ws.cell(row=row, column=2, value=val).font = NOTE_FONT
        row += 1
    
    # 岗位分布
    row += 1
    ws.cell(row=row, column=1, value="岗位分布").font = Font(name="微软雅黑", size=12, bold=True)
    row += 1
    
    # 表头
    for col, header in enumerate(["岗位类别", "Primary", "Reserve", "合计"], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1
    
    # 统计
    cat_primary = Counter(r["category"] for r in primary)
    cat_reserve = Counter(r["category"] for r in reserve)
    all_cats = sorted(set(list(cat_primary.keys()) + list(cat_reserve.keys())), 
                      key=lambda c: -cat_primary.get(c, 0) - cat_reserve.get(c, 0))
    
    for cat in all_cats:
        ws.cell(row=row, column=1, value=cat).font = NOTE_FONT
        ws.cell(row=row, column=2, value=cat_primary.get(cat, 0)).font = NOTE_FONT
        ws.cell(row=row, column=3, value=cat_reserve.get(cat, 0)).font = NOTE_FONT
        ws.cell(row=row, column=4, value=cat_primary.get(cat, 0) + cat_reserve.get(cat, 0)).font = NOTE_FONT
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1
    
    # 合计行
    ws.cell(row=row, column=1, value="合计").font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=row, column=2, value=len(primary)).font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=row, column=3, value=len(reserve)).font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=row, column=4, value=len(primary) + len(reserve)).font = Font(name="微软雅黑", size=10, bold=True)
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = THIN_BORDER


# ─── 泄漏检查 ──────────────────────────────────────────────────────────

def check_leakage(records):
    """检查是否存在模型预测或旧 gold 泄漏"""
    forbidden_patterns = [
        "current_gold", "gold_title", "predicted_title",
        "model_prediction", "LLM prediction", "llm_prediction",
        "auto_label", "自动标签", "model_output", "predicted",
        "gold_skills", "gold_experience", "gold_education",
        "gold_duties", "gold_", "extracted_",
    ]
    
    issues = []
    for r in records:
        for key in r.keys():
            key_lower = key.lower()
            for pattern in forbidden_patterns:
                if pattern in key_lower:
                    issues.append(f"  ⚠️ sample_id={r.get('sample_id', '?')}: 字段 '{key}' 疑似泄漏 (匹配 '{pattern}')")
    
    # 检查人工标注字段是否全部为空
    filled_fields = []
    for r in records:
        for field in ANNOTATION_FIELDS:
            val = r.get(field, "")
            if val and str(val).strip():
                filled_fields.append(f"  ⚠️ sample_id={r.get('sample_id', '?')}: '{field}' = '{str(val)[:50]}'")
    
    return issues, filled_fields


# ─── 主流程 ────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  人工标注数据包准备")
    print(f"  随机种子: {SEED}")
    print("=" * 60)
    
    # 1. 加载数据
    print("\n[1] 加载 accepted 记录...")
    accepted = load_accepted()
    print(f"  accepted: {len(accepted)} 条")
    
    # 验证数量
    assert len(accepted) == 135, f"Expected 135 accepted, got {len(accepted)}"
    
    # 2. 分层抽样
    print("\n[2] 确定性分层抽样...")
    primary, reserve = stratified_sample(accepted)
    print(f"  primary: {len(primary)} 条")
    print(f"  reserve: {len(reserve)} 条")
    assert len(primary) == 110, f"Expected 110 primary, got {len(primary)}"
    assert len(reserve) == 25, f"Expected 25 reserve, got {len(reserve)}"
    
    # 3. 检查重复 (sample_id)
    primary_ids = set(r["sample_id"] for r in primary)
    reserve_ids = set(r["sample_id"] for r in reserve)
    overlap = primary_ids & reserve_ids
    assert len(overlap) == 0, f"Primary and reserve sample_id overlap: {overlap}"
    
    # 3b. 检查重复 (source_id)
    primary_source_ids = set(r["source_id"] for r in primary)
    reserve_source_ids = set(r["source_id"] for r in reserve)
    source_id_overlap = primary_source_ids & reserve_source_ids
    assert len(source_id_overlap) == 0, f"Primary and reserve source_id overlap: {source_id_overlap}"
    
    # 4. 岗位分布
    print("\n[3] Primary 岗位分布:")
    cat_dist = Counter(r["category"] for r in primary)
    for cat, cnt in sorted(cat_dist.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {cnt}")
    
    # 5. 创建输出目录
    os.makedirs(ANNOTATION_DIR, exist_ok=True)
    
    # 6. 生成主标注 Excel
    print("\n[4] 生成主标注 Excel...")
    wb = openpyxl.Workbook()
    
    # Sheet 1: Round1盲标
    ws1 = wb.active
    create_annotation_sheet(ws1, primary, "Round1盲标")
    
    # Sheet 2: 标注说明
    ws2 = wb.create_sheet()
    create_instruction_sheet(ws2)
    
    # Sheet 3: 数据统计
    ws3 = wb.create_sheet()
    create_stats_sheet(ws3, primary, reserve)
    
    primary_path = os.path.join(ANNOTATION_DIR, "jd_annotation_round1_110.xlsx")
    wb.save(primary_path)
    print(f"  ✓ {primary_path}")
    
    # 7. 生成备用集 Excel
    print("\n[5] 生成备用集 Excel...")
    wb2 = openpyxl.Workbook()
    ws_reserve = wb2.active
    create_annotation_sheet(ws_reserve, reserve, "备用集")
    
    reserve_path = os.path.join(ANNOTATION_DIR, "jd_annotation_reserve_25.xlsx")
    wb2.save(reserve_path)
    print(f"  ✓ {reserve_path}")
    
    # 8. 生成样本清单 CSV
    print("\n[6] 生成样本清单 CSV...")
    manifest_path = os.path.join(ANNOTATION_DIR, "annotation_sampling_manifest.csv")
    all_records = primary + reserve
    with open(manifest_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["sample_id", "source_id", "job_title_raw", "job_category", "annotation_split"])
        writer.writeheader()
        for r in all_records:
            writer.writerow({
                "sample_id": r["sample_id"],
                "source_id": r["source_id"],
                "job_title_raw": r.get("job_title_raw", ""),
                "job_category": r.get("category", ""),
                "annotation_split": r["annotation_split"],
            })
    print(f"  ✓ {manifest_path} ({len(all_records)} 条)")
    
    # 9. 泄漏检查
    print("\n[7] 泄漏检查...")
    primary_issues, primary_filled = check_leakage(primary)
    reserve_issues, reserve_filled = check_leakage(reserve)
    
    all_issues = primary_issues + reserve_issues
    all_filled = primary_filled + reserve_filled
    
    if all_issues:
        print("  ❌ 发现泄漏字段:")
        for issue in all_issues:
            print(issue)
    else:
        print("  ✅ 无模型预测 / 旧 gold 泄漏")
    
    if all_filled:
        print("  ❌ 人工标注字段被预填:")
        for f in all_filled:
            print(f)
    else:
        print("  ✅ 人工标注字段全部为空")
    
    # 10. 最终报告
    print("\n" + "=" * 60)
    print("  准备完成!")
    print(f"  Primary: {len(primary)} 条 → {primary_path}")
    print(f"  Reserve: {len(reserve)} 条 → {reserve_path}")
    print(f"  Manifest: {len(all_records)} 条 → {manifest_path}")
    print(f"  泄漏检查: {'❌ 有问题' if all_issues or all_filled else '✅ 通过'}")
    print("=" * 60)


if __name__ == "__main__":
    main()